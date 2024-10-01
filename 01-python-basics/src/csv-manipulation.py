import openpyxl
from openpyxl.styles import PatternFill, Font
import os
file_path = os.path.join(os.path.dirname(__file__), 'xlsx', 'inventory.xlsx')
new_file_path = os.path.join(os.path.dirname(__file__), 'xlsx', 'inventory_added.xlsx')
workbook = openpyxl.load_workbook(file_path)
main_sheet = workbook['Sheet1']
PRODUCT_NO_COL = 1
INVENTORY_COL = 2
PRICE_COL = 3
COMPANY_NAME_COL = 4
NEW_INVENTORY_VALUE_COL = 5

company_product_cnt = {}
company_inventory_value = {}
products_low_stock = []

for rows in range (2, main_sheet.max_row +1):
  product_no_value = main_sheet.cell(rows, PRODUCT_NO_COL).value
  inventory_value = main_sheet.cell(rows, INVENTORY_COL).value
  price_value = main_sheet.cell(rows, PRICE_COL).value
  company_name_value = main_sheet.cell(rows, COMPANY_NAME_COL).value
  # count products per company
  if company_name_value not in company_product_cnt:
    company_product_cnt[company_name_value] = 1
  else:
    company_product_cnt[company_name_value] += 1
  # list all product no with low stock
  if inventory_value < 10:
    products_low_stock.append({"product_no": int(product_no_value), "inventory": int(inventory_value)})
  # calculate inventory value of company
  if company_name_value not in company_inventory_value:
    company_inventory_value[company_name_value] = inventory_value * price_value
  else:
    company_inventory_value[company_name_value] += inventory_value * price_value
  # create new column and fill with inventory value
  main_sheet.cell(rows, NEW_INVENTORY_VALUE_COL).value = inventory_value * price_value

# add new header
main_sheet.cell(1, NEW_INVENTORY_VALUE_COL).value = 'Inventory Value'
grey_fill = PatternFill(start_color="efefef", end_color="efefef", fill_type="solid")
bold_font = Font(name="Arial", size=11, bold=True)
main_sheet.cell(1, NEW_INVENTORY_VALUE_COL).fill = grey_fill
main_sheet.cell(1, NEW_INVENTORY_VALUE_COL).font = bold_font
workbook.save(new_file_path)

print(f"\n\n---------------------------------[Products with low stock]-------------------------------------------------------")
print(f"{products_low_stock}\n")
print(f"---------------------------------[Distinct products per company]-------------------------------------------------")
print(f"{company_product_cnt}\n")
print(f"---------------------------------[Value of company's inventory]--------------------------------------------------")
print(f"{company_inventory_value}\n")

print(f" ___________________________________________________________\n" \
      f"|                                                           |\n"\
      f"|             new column Inventory Value persisted          |\n"\
      f"|          xlsx file created in xlsx/inventory_added.xlsx   |\n"\
      f"|                                                           |\n"\
      f"|___________________________________________________________|" \
      )